import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
from openpyxl import load_workbook

def getKey(dic, valor):
    for c, v in dic.items():
        if v == valor:
            return c

def getCities(sheet, linha, coluna):
    cidades = {}

    cell = sheet.cell(row=linha, column=coluna).value
    while cell is not None:
        cidades[coluna] = cell

        coluna += 1
        cell = sheet.cell(row=linha, column=coluna).value

    return cidades

def main():
    st.header("Painel de Acompanhamento, Monitoramento e Avaliação dos Indicadores da Saúde do Estado")
    st.subheader("Diretoria de Unidade de Planejamento (DUP)")
    st.divider()
    planilha = st.file_uploader('Selecione a planilha', type='xlsx')
    st.divider()

    if planilha is not None:
        workbook = load_workbook(planilha)

        dataDictionary = {"MACROPROBLEMA":[], "CIDADE":[], "AVALIAÇÃO":[], "TOTAL":[], "PRIORIDADE":[], "MACROREGIAO":[]}
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]

            cidades = getCities(sheet, linha = 2, coluna = 2)

            # Constrói data dictionary
            linha = 3
            macroProblema = sheet.cell(row=linha, column=1).value
            while macroProblema is not None:
                coluna = 2
                total = sheet.cell(row=linha, column=getKey(cidades,"TOTAL")).value
                priodade = sheet.cell(row=linha, column=getKey(cidades,"PRIORIDADES SANITÁRIAS")).value
                avaliacao = sheet.cell(row=linha, column=coluna).value
                while avaliacao is not None and cidades[coluna] != "TOTAL":

                    dataDictionary["MACROPROBLEMA"].append(macroProblema)
                    dataDictionary["CIDADE"].append(cidades[coluna])
                    dataDictionary["AVALIAÇÃO"].append(avaliacao)
                    dataDictionary["TOTAL"].append(total)
                    dataDictionary["PRIORIDADE"].append(priodade)
                    dataDictionary["MACROREGIAO"].append(sheet.title)

                    coluna += 1
                    avaliacao = sheet.cell(row=linha, column=coluna).value

                linha += 1
                macroProblema = sheet.cell(row=linha, column=1).value

        df = pd.DataFrame(dataDictionary)

        macroRegiaoOptions = np.append(['TODAS'], df["MACROREGIAO"].unique())
        macroRegiaoEscolhida = st.selectbox(label="Macro Região", options=macroRegiaoOptions)

        if macroRegiaoEscolhida == 'TODAS':
            macroProblemaOptions = np.append(['TODOS'], df["MACROPROBLEMA"].unique())
            macroProblemaEscolhido = st.selectbox(label="Macro Problema", options=macroProblemaOptions)
            # st.table(macroProblemaOptions)
        else:
            macroProblemaDaMacroRegiao = np.append(['TODOS'], df.loc[(df['MACROREGIAO'] == macroRegiaoEscolhida)]["MACROPROBLEMA"].unique())
            macroProblemaEscolhido = st.selectbox(label="Macro Problema", options=macroProblemaDaMacroRegiao)
            # st.table(macroProblemaDaMacroRegiao)


        if macroProblemaEscolhido:
            if macroProblemaEscolhido == 'TODOS':
                if macroRegiaoEscolhida == 'TODAS':
                    df_filtered = df
                else:
                    df_filtered = df.loc[(df["MACROREGIAO"] == macroRegiaoEscolhida)]
            else:
                df_filtered = df.loc[(df["MACROPROBLEMA"] == macroProblemaEscolhido)]
            
            
            bar = px.bar(df_filtered, x='CIDADE', y ='AVALIAÇÃO', color='MACROPROBLEMA')
            bar.update_traces(showlegend=False, hoverlabel_align='left', selector=dict(type='bar'))
            bar.update_layout(xaxis={'categoryorder':'total descending'})
            st.plotly_chart(bar, use_container_width=True)


main()